import streamlit as st
import google.generativeai as genai
import os
from dotenv import load_dotenv
from PyPDF2 import PdfReader
import docx
import openpyxl

# Load API key từ environment (Streamlit Cloud sẽ set sau)
load_dotenv()
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

st.title("✈️ Aircraft Maintenance AI Assistant")

uploaded_file = st.file_uploader("📂 Upload a document (PDF, Word, Excel)", type=["pdf", "docx", "xlsx"])

def read_file(file):
    text = ""
    if file.name.endswith(".pdf"):
        reader = PdfReader(file)
        for page in reader.pages:
            text += page.extract_text()
    elif file.name.endswith(".docx"):
        doc = docx.Document(file)
        for para in doc.paragraphs:
            text += para.text + "\n"
    elif file.name.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text

if uploaded_file:
    st.success("✅ File uploaded successfully")
    content = read_file(uploaded_file)
    st.write("📄 Extracted text preview:", content[:1000])  # show 1000 chars

    question = st.text_input("❓ Ask a question about this document:")

    if st.button("Analyze with Gemini"):
        if question:
            model = genai.GenerativeModel("gemini-pro")
            response = model.generate_content([content, question])
            st.subheader("💡 AI Answer:")
            st.write(response.text)
        else:
            st.warning("⚠️ Please enter a question.")
